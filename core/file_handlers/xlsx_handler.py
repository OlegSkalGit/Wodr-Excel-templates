import os
import re
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
from core.excel_styles import deserialize_color, deserialize_fill, deserialize_alignment, deserialize_border

def apply_cell_styles(cell, key, variables):
    from copy import copy
    # 1. Bold
    bold_key = f"{key}_bold"
    if bold_key in variables:
        val = variables[bold_key]
        if val != "" and val is not None:
            is_bold = str(val).lower() in ["true", "1", "yes"]
            if cell.font:
                new_font = copy(cell.font)
                new_font.bold = is_bold
                cell.font = new_font
            else:
                from openpyxl.styles import Font
                cell.font = Font(bold=is_bold)
                
    # 2. Italic
    italic_key = f"{key}_italic"
    if italic_key in variables:
        val = variables[italic_key]
        if val != "" and val is not None:
            is_italic = str(val).lower() in ["true", "1", "yes"]
            if cell.font:
                new_font = copy(cell.font)
                new_font.italic = is_italic
                cell.font = new_font
            else:
                from openpyxl.styles import Font
                cell.font = Font(italic=is_italic)
                
    # 3. Font Size
    size_key = f"{key}_font_size"
    if size_key in variables:
        val = variables[size_key]
        if val != "" and val is not None:
            try:
                sz_val = float(val)
                if cell.font:
                    new_font = copy(cell.font)
                    new_font.sz = sz_val
                    cell.font = new_font
                else:
                    from openpyxl.styles import Font
                    cell.font = Font(size=sz_val)
            except: pass
            
    # 4. Font Name
    name_key = f"{key}_font_name"
    if name_key in variables:
        val = variables[name_key]
        if val != "" and val is not None:
            if cell.font:
                new_font = copy(cell.font)
                new_font.name = str(val)
                cell.font = new_font
            else:
                from openpyxl.styles import Font
                cell.font = Font(name=str(val))
                
    # 5. Font Color
    color_key = f"{key}_font_color"
    if color_key in variables:
        val = variables[color_key]
        if val != "" and val is not None:
            color_obj = deserialize_color(str(val))
            if cell.font:
                new_font = copy(cell.font)
                new_font.color = color_obj
                cell.font = new_font
            else:
                from openpyxl.styles import Font
                cell.font = Font(color=color_obj)
                
    # 6. Fill
    fill_key = f"{key}_fill"
    if fill_key in variables:
        val = variables[fill_key]
        if val != "" and val is not None:
            fill_obj = deserialize_fill(str(val))
            cell.fill = fill_obj
            
    # 7. Alignment
    align_key = f"{key}_alignment"
    if align_key in variables:
        val = variables[align_key]
        if val != "" and val is not None:
            align_obj = deserialize_alignment(str(val))
            cell.alignment = align_obj
            
    # 8. Border
    border_key = f"{key}_border"
    if border_key in variables:
        val = variables[border_key]
        if val != "" and val is not None:
            border_obj = deserialize_border(str(val))
            cell.border = border_obj
            
    # 9. Number Format
    num_key = f"{key}_number_format"
    if num_key in variables:
        val = variables[num_key]
        if val != "" and val is not None:
            cell.number_format = str(val)

def process_excel(template_path, output_path, variables):
    try:
        wb = openpyxl.load_workbook(template_path)
        str_vars = {k: str(v) for k, v in variables.items()}
        for sheet in wb.worksheets:
            new_title = sheet.title
            title_modified = False
            for key, val in str_vars.items():
                pattern = r"\{\{\s*" + re.escape(key) + r"\s*\}\}"
                if re.search(pattern, new_title):
                    new_title = re.sub(pattern, lambda _: val, new_title)
                    title_modified = True
            if title_modified:
                clean_title = re.sub(r'[\\/*?:\[\]]', "", new_title)[:31].strip()
                if not clean_title: clean_title = "Sheet"
                base_title = clean_title
                counter = 1
                while clean_title in wb.sheetnames and clean_title != sheet.title:
                    suffix = f"_{counter}"
                    limit = 31 - len(suffix)
                    clean_title = base_title[:limit] + suffix
                    counter += 1
                sheet.title = clean_title
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        exact_key = None
                        for key in variables.keys():
                            pattern = r"^\s*\{\{\s*" + re.escape(key) + r"\s*\}\}\s*$"
                            if re.match(pattern, original_value):
                                exact_key = key
                                break
                        if exact_key is not None:
                            new_value = variables[exact_key]
                            source_type = variables.get(f"{exact_key}_type_code", "s")
                            template_type = cell.data_type
                            target_type = source_type if source_type != "s" else template_type
                            if isinstance(new_value, str):
                                if target_type == "n":
                                    try:
                                        if "." in new_value or "e" in new_value.lower(): new_value = float(new_value)
                                        else: new_value = int(new_value)
                                    except: pass
                                elif target_type == "d":
                                    try: new_value = datetime.fromisoformat(new_value)
                                    except: pass
                                elif target_type == "b":
                                    if new_value.lower() in ["true", "1", "yes"]: new_value = True
                                    elif new_value.lower() in ["false", "0", "no"]: new_value = False
                            cell.value = new_value
                            apply_cell_styles(cell, exact_key, variables)
                            if isinstance(cell.value, str) and '\n' in cell.value: cell.alignment = Alignment(wrapText=True)
                            continue
                        new_value = original_value
                        modified = False
                        matched_keys = []
                        for key, val in str_vars.items():
                            pattern = r"\{\{\s*" + re.escape(key) + r"\s*\}\}"
                            if re.search(pattern, new_value):
                                new_value = re.sub(pattern, lambda _: val, new_value)
                                modified = True
                                matched_keys.append(key)
                        if modified:
                            cell.value = new_value
                            for key in matched_keys:
                                apply_cell_styles(cell, key, variables)
                            if isinstance(new_value, str) and '\n' in new_value: cell.alignment = Alignment(wrapText=True)
        wb.save(output_path)
        return True, None
    except Exception as e:
        err = f"[Помилка Excel]: {e}"
        print(f"  {err}")
        return False, err
