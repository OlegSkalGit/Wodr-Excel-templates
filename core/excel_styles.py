import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color

def serialize_color(color_obj):
    if color_obj is None:
        return ""
    parts = []
    
    def is_valid(val):
        if val is None:
            return False
        # Check for openpyxl descriptor exception strings/classes
        if type(val).__name__ in ['Integer', 'Float', 'String', 'Bool', 'NoneSet', 'MinMax', 'Set', 'Alias']:
            return False
        val_str = str(val)
        if val_str.startswith("Values must be"):
            return False
        return True

    if is_valid(color_obj.type):
        parts.append(f"type={color_obj.type}")
    if is_valid(color_obj.rgb):
        parts.append(f"rgb={color_obj.rgb}")
    if is_valid(color_obj.theme):
        parts.append(f"theme={color_obj.theme}")
    if is_valid(color_obj.tint) and color_obj.tint != 0.0:
        parts.append(f"tint={color_obj.tint}")
    if is_valid(color_obj.indexed):
        parts.append(f"indexed={color_obj.indexed}")
    if is_valid(color_obj.auto):
        parts.append(f"auto={color_obj.auto}")
        
    return ",".join(parts)

def deserialize_color(color_str):
    if not color_str:
        return None
    kwargs = {}
    for part in color_str.split(","):
        if "=" in part:
            k, v = part.split("=", 1)
            if k in ['theme', 'indexed']:
                kwargs[k] = int(v)
            elif k == 'tint':
                kwargs[k] = float(v)
            elif k == 'auto':
                kwargs[k] = v.lower() == 'true'
            else:
                kwargs[k] = v
    return Color(**kwargs)

def serialize_fill(fill_obj):
    if fill_obj is None or fill_obj.fill_type is None:
        return ""
    parts = []
    parts.append(f"fill_type={fill_obj.fill_type}")
    if fill_obj.fgColor:
        fg = serialize_color(fill_obj.fgColor)
        if fg: parts.append(f"fgColor={fg}")
    if fill_obj.bgColor:
        bg = serialize_color(fill_obj.bgColor)
        if bg: parts.append(f"bgColor={bg}")
    return ";".join(parts)

def deserialize_fill(fill_str):
    if not fill_str:
        return PatternFill(fill_type=None)
    kwargs = {}
    for part in fill_str.split(";"):
        if "=" in part:
            k, v = part.split("=", 1)
            if k == 'fill_type':
                kwargs['fill_type'] = v
            elif k == 'fgColor':
                kwargs['fgColor'] = deserialize_color(v)
            elif k == 'bgColor':
                kwargs['bgColor'] = deserialize_color(v)
    return PatternFill(**kwargs)

def serialize_alignment(align_obj):
    if align_obj is None:
        return ""
    parts = []
    
    def is_valid(val):
        if val is None:
            return False
        if type(val).__name__ in ['Integer', 'Float', 'String', 'Bool', 'NoneSet', 'MinMax', 'Set', 'Alias']:
            return False
        val_str = str(val)
        if val_str.startswith("Values must be"):
            return False
        return True

    if is_valid(align_obj.horizontal):
        parts.append(f"horizontal={align_obj.horizontal}")
    if is_valid(align_obj.vertical):
        parts.append(f"vertical={align_obj.vertical}")
    if is_valid(align_obj.text_rotation):
        parts.append(f"text_rotation={align_obj.text_rotation}")
    if is_valid(align_obj.wrap_text):
        parts.append(f"wrap_text={align_obj.wrap_text}")
    if is_valid(align_obj.shrink_to_fit):
        parts.append(f"shrink_to_fit={align_obj.shrink_to_fit}")
    if is_valid(align_obj.indent):
        parts.append(f"indent={align_obj.indent}")
    return ",".join(parts)

def deserialize_alignment(align_str):
    if not align_str:
        return None
    kwargs = {}
    for part in align_str.split(","):
        if "=" in part:
            k, v = part.split("=", 1)
            if k in ['text_rotation', 'indent']:
                kwargs[k] = int(float(v))
            elif k in ['wrap_text', 'shrink_to_fit']:
                kwargs[k] = v.lower() == 'true'
            else:
                kwargs[k] = v
    return Alignment(**kwargs)

def serialize_side(side_obj):
    if side_obj is None or side_obj.style is None:
        return ""
    parts = []
    parts.append(f"style={side_obj.style}")
    if side_obj.color:
        col = serialize_color(side_obj.color)
        if col: parts.append(f"color={col}")
    return ";".join(parts)

def deserialize_side(side_str):
    if not side_str:
        return None
    kwargs = {}
    for part in side_str.split(";"):
        if "=" in part:
            k, v = part.split("=", 1)
            if k == 'style':
                kwargs['style'] = v
            elif k == 'color':
                kwargs['color'] = deserialize_color(v)
    return Side(**kwargs)

def serialize_border(border_obj):
    if border_obj is None:
        return ""
    parts = []
    for side_name in ['left', 'right', 'top', 'bottom']:
        side = getattr(border_obj, side_name, None)
        side_str = serialize_side(side)
        if side_str:
            parts.append(f"{side_name}={side_str}")
    return "|".join(parts)

def deserialize_border(border_str):
    if not border_str:
        return None
    kwargs = {}
    for part in border_str.split("|"):
        if "=" in part:
            side_name, side_str = part.split("=", 1)
            kwargs[side_name] = deserialize_side(side_str)
    return Border(**kwargs)

def check_cell_style_diffs(sheets_d, r, c):
    cells = [s.cell(row=r, column=c) for s in sheets_d]
    diffs = {}
    
    # 1. Bold
    bolds = [bool(cell.font.bold) if cell.font else False for cell in cells]
    if len(set(bolds)) > 1:
        diffs['bold'] = bolds
        
    # 2. Italic
    italics = [bool(cell.font.italic) if cell.font else False for cell in cells]
    if len(set(italics)) > 1:
        diffs['italic'] = italics
        
    # 3. Font size
    sizes = [float(cell.font.size) if (cell.font and cell.font.size) else 11.0 for cell in cells]
    if len(set(sizes)) > 1:
        diffs['font_size'] = sizes
        
    # 4. Font name
    names = [str(cell.font.name) if (cell.font and cell.font.name) else "Calibri" for cell in cells]
    if len(set(names)) > 1:
        diffs['font_name'] = names
        
    # 5. Font color
    colors = [serialize_color(cell.font.color) if (cell.font and cell.font.color) else "" for cell in cells]
    if len(set(colors)) > 1:
        diffs['font_color'] = colors
        
    # 6. Fill
    fills = [serialize_fill(cell.fill) if cell.fill else "" for cell in cells]
    if len(set(fills)) > 1:
        diffs['fill'] = fills
        
    # 7. Alignment
    aligns = [serialize_alignment(cell.alignment) if cell.alignment else "" for cell in cells]
    if len(set(aligns)) > 1:
        diffs['alignment'] = aligns
        
    # 8. Border
    borders = [serialize_border(cell.border) if cell.border else "" for cell in cells]
    if len(set(borders)) > 1:
        diffs['border'] = borders
        
    # 9. Number format
    num_formats = [str(cell.number_format or "General") for cell in cells]
    if len(set(num_formats)) > 1:
        diffs['number_format'] = num_formats
        
    return diffs
