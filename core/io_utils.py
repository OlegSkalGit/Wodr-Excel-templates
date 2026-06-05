import os
import re
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

def is_excel_config(file_path):
    """Robust algorithm to distinguish config Excel files from regular spreadsheets."""
    if not file_path.endswith('.xlsx') or os.path.basename(file_path).startswith('~$'):
        return False
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Read cells A1 and A2
            a1 = ws.cell(row=1, column=1).value
            a2 = ws.cell(row=2, column=1).value
            if a1 and isinstance(a1, str):
                a1_lower = a1.lower()
                if any(ext in a1_lower for ext in ['.docx', '.xlsx']) or 'template' in a1_lower:
                    return True
            if a2 and isinstance(a2, str):
                a2_lower = a2.lower()
                if any(ext in a2_lower for ext in ['.docx', '.xlsx']) or '{{' in a2_lower:
                    return True
        return False
    except Exception:
        return False

def scan_recursive_configs(root_folder):
    """Walks the folder structure recursively and returns a list of config files."""
    config_files = []
    if not os.path.exists(root_folder) or not os.path.isdir(root_folder):
        return []
    for dirpath, _, filenames in os.walk(root_folder):
        for f in filenames:
            if f.endswith('.xlsx') and not f.startswith('~$'):
                full_path = os.path.abspath(os.path.join(dirpath, f))
                if is_excel_config(full_path):
                    config_files.append(full_path)
    return sorted(config_files)

def build_dir_tree(config_files, root_path):
    """Builds a nested directory tree dictionary from config paths."""
    tree = {}
    for path in config_files:
        rel_path = os.path.relpath(path, root_path)
        parts = rel_path.split(os.sep)
        current = tree
        for part in parts[:-1]:
            if part not in current:
                current[part] = {}
            current = current[part]
        current[parts[-1]] = path
    return tree

def build_docs_only_tree(config_files, root_path, config_loader=None, active_vars_loader=None):
    """Builds a virtual tree where documents are grouped directly under folder structures,
    skipping config files and sheets nodes."""
    from core.text_processor import resolve_virtual_doc_name
    
    if config_loader is None:
        config_loader = load_excel_config
        
    tree = {}
    abs_root_path = os.path.abspath(root_path)
    
    for path in config_files:
        abs_config_path = os.path.abspath(path)
        config_dir = os.path.dirname(abs_config_path)
            
        sheets_data = config_loader(abs_config_path)
        if not sheets_data:
            continue
            
        for sheet_name, info in sheets_data.items():
            rows = info.get("rows", [])
            template_path = info.get("template_path", "")
            name_pattern = info.get("name_pattern", "")
            
            for idx, row in enumerate(rows):
                # Try getting the current active edits if provided by caller (for preview updating)
                current_row_vars = row
                if active_vars_loader:
                    current_row_vars = active_vars_loader(abs_config_path, sheet_name, idx, row)
                
                doc_name = resolve_virtual_doc_name(name_pattern, current_row_vars, template_path)
                if not doc_name.strip():
                    doc_name = f"document_{idx + 1}"
                
                # Resolve absolute path of the document
                if os.path.isabs(doc_name):
                    abs_doc_path = os.path.abspath(doc_name)
                else:
                    abs_doc_path = os.path.abspath(os.path.join(config_dir, doc_name))
                    
                abs_doc_dir = os.path.dirname(abs_doc_path)
                
                # Determine relative folder directory to root_path
                try:
                    rel_doc_dir = os.path.relpath(abs_doc_dir, abs_root_path)
                    norm_rel = rel_doc_dir.replace("\\", "/")
                    rel_parts = [p for p in norm_rel.split("/") if p]
                    # If it goes outside the root (starts with '..') or is absolute, use absolute path
                    if (rel_parts and rel_parts[0] == "..") or os.path.isabs(rel_doc_dir):
                        rel_doc_dir = abs_doc_dir
                except ValueError:
                    rel_doc_dir = abs_doc_dir
                
                # Normalize separators for path splitting
                rel_doc_dir = rel_doc_dir.replace("\\", "/")
                
                if rel_doc_dir == "." or not rel_doc_dir:
                    parts = []
                else:
                    parts = rel_doc_dir.split("/")
                    parts = [p for p in parts if p and p != "."]
                    
                doc_info = {
                    "config_path": abs_config_path,
                    "sheet_name": sheet_name,
                    "row_idx": idx,
                    "doc_name": doc_name,
                    "template_path": template_path,
                    "name_pattern": name_pattern
                }
                
                current = tree
                for part in parts:
                    if part not in current:
                        current[part] = {}
                    current = current[part]
                
                if "__docs__" not in current:
                    current["__docs__"] = []
                current["__docs__"].append(doc_info)
                
    return tree

def load_excel_config(filepath):
    """Loads all worksheets from Excel config safely parsing metadata (A1 and A2)."""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Файл не знайдено: {filepath}")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_data = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Template is in Cell A1
        template_path = sheet.cell(row=1, column=1).value
        # Output file name pattern is in Cell A2
        name_pattern = sheet.cell(row=2, column=1).value
        
        if template_path is not None:
            template_path = str(template_path).strip()
        else:
            template_path = ""
            
        if name_pattern is not None:
            name_pattern = str(name_pattern).strip()
        else:
            name_pattern = ""
            
        # Headers are located on Row 4
        headers = []
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=4, column=col).value
            if val is not None:
                header_name = str(val).strip()
                if header_name.startswith("{{") and header_name.endswith("}}"):
                    header_name = header_name[2:-2].strip()
                headers.append(header_name)
            else:
                headers.append("")
        
        # Remove trailing empty headers
        while headers and headers[-1] == "":
            headers.pop()
            
        if not headers:
            continue
            
        # Data rows start from Row 5
        rows = []
        for r in range(5, sheet.max_row + 1):
            row_vals = {}
            has_value = False
            for col_idx, h in enumerate(headers):
                cell = sheet.cell(row=r, column=col_idx + 1)
                cell_val = cell.value
                if cell_val is not None:
                    has_value = True
                row_vals[h] = str(cell_val) if cell_val is not None else ""
                row_vals[f"{h}_type_code"] = cell.data_type
            if has_value:
                rows.append(row_vals)
                
        sheets_data[sheet_name] = {
            "template_path": template_path,
            "name_pattern": name_pattern,
            "headers": headers,
            "rows": rows
        }
    return sheets_data

def save_excel_config(filepath, sheet_name, template_path, name_pattern, headers, df_data):
    """Saves changes back to Excel config safely preserving other sheets."""
    config_dir = os.path.dirname(os.path.abspath(filepath))
    try:
        if os.path.isabs(template_path):
            target_abs = os.path.abspath(template_path)
            if os.path.splitdrive(target_abs)[0].lower() == os.path.splitdrive(config_dir)[0].lower():
                template_path = os.path.relpath(target_abs, config_dir)
    except Exception:
        pass

    wb = openpyxl.load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Аркуш '{sheet_name}' не знайдено у файлі конфігурації.")
        
    sheet = wb[sheet_name]
    
    # 1. Update metadata in A1 and A2
    sheet.cell(row=1, column=1).value = template_path
    sheet.cell(row=2, column=1).value = name_pattern
    
    # 2. Update headers in Row 4 and clear deleted ones
    max_c = max(sheet.max_column, len(headers))
    for c_idx in range(1, max_c + 1):
        if c_idx <= len(headers):
            sheet.cell(row=4, column=c_idx).value = headers[c_idx - 1]
        else:
            sheet.cell(row=4, column=c_idx).value = None
        
    # 3. Clear old data rows from Row 5 onwards up to max_column
    max_r = max(sheet.max_row, 5)
    for r in range(5, max_r + 1):
        for c in range(1, max_c + 1):
            sheet.cell(row=r, column=c).value = None
            
    # 4. Write new data rows
    for r_idx, row_dict in enumerate(df_data):
        for c_idx, h in enumerate(headers):
            val = row_dict.get(h, "")
            cell = sheet.cell(row=5 + r_idx, column=c_idx + 1)
            
            type_code = row_dict.get(f"{h}_type_code", "s")
            
            if isinstance(val, str):
                val_strip = val.strip()
                if type_code == "n":
                    try:
                        if "." in val_strip or "e" in val_strip.lower():
                            cell.value = float(val_strip)
                        else:
                            cell.value = int(val_strip)
                    except ValueError:
                        cell.value = val
                elif type_code == "b":
                    if val_strip.lower() in ["true", "1", "yes"]:
                        cell.value = True
                    elif val_strip.lower() in ["false", "0", "no"]:
                        cell.value = False
                    else:
                        cell.value = val
                elif type_code == "d":
                    try:
                        cell.value = datetime.fromisoformat(val_strip)
                    except ValueError:
                        cell.value = val
                else:
                    if val_strip.isdigit():
                        cell.value = int(val_strip)
                    elif re.match(r'^\d+\.\d+$', val_strip):
                        cell.value = float(val_strip)
                    else:
                        cell.value = val
            else:
                cell.value = val
                
            # Enable multiline wrapping if newlines are present
            if isinstance(cell.value, str) and '\n' in cell.value:
                cell.alignment = Alignment(wrapText=True)
                
    wb.save(filepath)
    return True

def update_config_template_path(filepath, sheet_name, new_template_path):
    """Safely updates only the template path (cell A1) of a specific sheet in Excel config."""
    try:
        config_dir = os.path.dirname(os.path.abspath(filepath))
        try:
            if os.path.isabs(new_template_path):
                target_abs = os.path.abspath(new_template_path)
                if os.path.splitdrive(target_abs)[0].lower() == os.path.splitdrive(config_dir)[0].lower():
                    new_template_path = os.path.relpath(target_abs, config_dir)
        except Exception:
            pass
            
        wb = openpyxl.load_workbook(filepath)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.cell(row=1, column=1).value = new_template_path
            wb.save(filepath)
            return True, None
        return False, f"Аркуш '{sheet_name}' не знайдено."
    except Exception as e:
        return False, str(e)

def create_new_sheet(filepath, new_sheet_name):
    """Creates a new worksheet with default metadata and template placeholders."""
    try:
        wb = openpyxl.load_workbook(filepath)
        if new_sheet_name in wb.sheetnames:
            return False, f"Аркуш з назвою '{new_sheet_name}' вже існує!"
        ws = wb.create_sheet(title=new_sheet_name)
        # Initialize default values
        ws['A1'] = "template_placeholder.docx"
        ws['A2'] = "output_{{YYYY}}{{MM}}{{DD}}.docx"
        ws.cell(row=4, column=1).value = "field_1"
        ws.cell(row=5, column=1).value = "значення_1"
        wb.save(filepath)
        return True, None
    except Exception as e:
        return False, str(e)

def rename_sheet(filepath, old_sheet_name, new_sheet_name):
    """Safely renames a sheet in the Excel config."""
    try:
        wb = openpyxl.load_workbook(filepath)
        if old_sheet_name not in wb.sheetnames:
            return False, f"Аркуш '{old_sheet_name}' не знайдено!"
        if new_sheet_name in wb.sheetnames:
            return False, f"Аркуш з назвою '{new_sheet_name}' вже існує!"
        ws = wb[old_sheet_name]
        ws.title = new_sheet_name
        wb.save(filepath)
        return True, None
    except Exception as e:
        return False, str(e)

def delete_sheet(filepath, sheet_name):
    """Deletes a sheet from the Excel config."""
    try:
        wb = openpyxl.load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return False, f"Аркуш '{sheet_name}' не знайдено!"
        if len(wb.sheetnames) <= 1:
            return False, "Неможливо видалити єдиний аркуш у книзі!"
        wb.remove(wb[sheet_name])
        wb.save(filepath)
        return True, None
    except Exception as e:
        return False, str(e)
