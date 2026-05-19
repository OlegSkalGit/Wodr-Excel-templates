import streamlit as st
import os
import sys
import re
import openpyxl
import pandas as pd
import subprocess
import time

# Configure page layout and style
st.set_page_config(
    page_title="TemplateMachine Control Center",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom premium styling
st.markdown("""
<style>
    /* Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Outfit', sans-serif;
    }
    
    code, pre {
        font-family: 'JetBrains Mono', monospace !important;
    }
    
    /* Elegant Title and Badges */
    .main-title {
        background: linear-gradient(135deg, #4A90E2 0%, #50E3C2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 3rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
        letter-spacing: -0.5px;
    }
    .subtitle {
        color: #8c96a3;
        font-size: 1.15rem;
        margin-bottom: 2rem;
        font-weight: 300;
    }
    
    /* Premium Styled Card Container */
    .card {
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        border-radius: 16px;
        box-shadow: 0 8px 32px 0 rgba(31, 38, 135, 0.05);
        padding: 1.8rem;
        margin-bottom: 1.8rem;
        border: 1px solid rgba(255, 255, 255, 0.18);
    }
    
    /* Styled widgets & alerts */
    .stAlert {
        border-radius: 12px !important;
        border: none !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.02) !important;
    }
    
    /* Sidebar premium background */
    [data-testid="stSidebar"] {
        background-color: #f7f9fc !important;
        border-right: 1px solid #eef2f6 !important;
    }
    
    /* Button custom hover effects */
    button[kind="primary"] {
        background: linear-gradient(135deg, #4A90E2 0%, #357ABD 100%) !important;
        border: none !important;
        color: white !important;
        box-shadow: 0 4px 15px rgba(74, 144, 226, 0.3) !important;
        transition: all 0.25s ease-in-out !important;
    }
    button[kind="primary"]:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(74, 144, 226, 0.4) !important;
    }
    
    button[kind="secondary"] {
        border-radius: 8px !important;
        border: 1px solid #e2e8f0 !important;
        transition: all 0.2s ease !important;
    }
    button[kind="secondary"]:hover {
        border-color: #4A90E2 !important;
        color: #4A90E2 !important;
        background-color: rgba(74, 144, 226, 0.03) !important;
    }
    
    .badge-icon {
        font-size: 1.5rem;
        margin-right: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------
# HELPER FUNCTIONS FOR FILE & WINDOWS DIALOGS
# ----------------------------------------------------

def open_folder_picker(title="Оберіть папку"):
    """Opens a native Windows directory selection dialog."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        folder = filedialog.askdirectory(parent=root, title=title)
        root.destroy()
        return folder
    except Exception as e:
        st.warning("Не вдалося відкрити діалогове вікно Windows. Будь ласка, введіть шлях вручну.")
        return ""

def open_file_picker(filetypes=None):
    """Opens a native Windows file selection dialog."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        if filetypes is None:
            filetypes = [("Усі підтримувані", "*.docx;*.xlsx"), ("Word файли", "*.docx"), ("Excel файли", "*.xlsx"), ("Усі файли", "*.*")]
        file = filedialog.askopenfilename(parent=root, title="Оберіть файл-зразок", filetypes=filetypes)
        root.destroy()
        return file
    except Exception as e:
        st.warning("Не вдалося відкрити діалогове вікно Windows. Будь ласка, введіть шлях вручну.")
        return ""

def create_new_sheet(filepath, new_sheet_name):
    """Creates a new worksheet with default metadata and template placeholders."""
    try:
        wb = openpyxl.load_workbook(filepath)
        if new_sheet_name in wb.sheetnames:
            st.error(f"Аркуш з назвою '{new_sheet_name}' вже існує!")
            return False
        ws = wb.create_sheet(title=new_sheet_name)
        # Initialize default values
        ws['A1'] = "template_placeholder.docx"
        ws['A2'] = "output_{{YYYY}}{{MM}}{{DD}}.docx"
        ws.cell(row=4, column=1).value = "field_1"
        ws.cell(row=5, column=1).value = "значення_1"
        wb.save(filepath)
        return True
    except Exception as e:
        st.error(f"Помилка при створенні аркуша: {e}")
        return False

def rename_sheet(filepath, old_sheet_name, new_sheet_name):
    """Safely renames a sheet in the Excel config."""
    try:
        wb = openpyxl.load_workbook(filepath)
        if old_sheet_name not in wb.sheetnames:
            st.error(f"Аркуш '{old_sheet_name}' не знайдено!")
            return False
        if new_sheet_name in wb.sheetnames:
            st.error(f"Аркуш з назвою '{new_sheet_name}' вже існує!")
            return False
        ws = wb[old_sheet_name]
        ws.title = new_sheet_name
        wb.save(filepath)
        return True
    except Exception as e:
        st.error(f"Помилка при перейменуванні аркуша: {e}")
        return False

def delete_sheet(filepath, sheet_name):
    """Deletes a sheet from the Excel config."""
    try:
        wb = openpyxl.load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            st.error(f"Аркуш '{sheet_name}' не знайдено!")
            return False
        if len(wb.sheetnames) <= 1:
            st.error("Неможливо видалити єдиний аркуш у книзі!")
            return False
        wb.remove(wb[sheet_name])
        wb.save(filepath)
        return True
    except Exception as e:
        st.error(f"Помилка при видаленні аркуша: {e}")
        return False

def update_config_template_path(filepath, sheet_name, new_template_path):
    """Safely updates only the template path (cell A1) of a specific sheet in Excel config."""
    try:
        wb = openpyxl.load_workbook(filepath)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.cell(row=1, column=1).value = new_template_path
            wb.save(filepath)
            return True
        return False
    except Exception as e:
        st.error(f"Помилка при оновленні шляху шаблону в конфігу: {e}")
        return False

def get_formatted_documentation_markdown():
    """Returns the full technical guide content formatted as premium Markdown, loaded dynamically from _templates_machine_.txt."""
    try:
        filepath = "_templates_machine_.txt"
        if os.path.exists(filepath):
            with open(filepath, "r", encoding="utf-8") as f:
                return f.read()
    except Exception as e:
        st.warning(f"Не вдалося завантажити файл довідки: {e}")
    
    # Fallback to absolute bare-bones in case the file gets deleted
    return "### 📖 Документація\nНе вдалося завантажити файл довідки `_templates_machine_.txt`. Будь ласка, переконайтеся, що файл існує у робочій папці."

# ----------------------------------------------------
# SYSTEM STATE & WORKSPACE SCANNING
# ----------------------------------------------------

def scan_workspace():
    """Scans the directory for configs and templates."""
    files = os.listdir('.')
    configs = [f for f in files if f.endswith('.xlsx') and not f.startswith('~$') and 'template' not in f.lower()]
    templates = [f for f in files if (f.endswith('.docx') or f.endswith('.xlsx')) and f.startswith('template_')]
    
    # Check inside example directory as well
    if os.path.exists('example'):
        for f in os.listdir('example'):
            if f.endswith('.xlsx') and not f.startswith('~$') and 'template' not in f.lower():
                configs.append(os.path.join('example', f))
            if (f.endswith('.docx') or f.endswith('.xlsx')) and f.startswith('template_'):
                templates.append(os.path.join('example', f))
                
    return sorted(configs), sorted(templates)

# ----------------------------------------------------
# EXCEL CONFIG READ/WRITE WITH OPENPYXL
# ----------------------------------------------------

def load_excel_config(filepath):
    """Loads all data sheets from an Excel config safely."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            template_path = sheet.cell(row=1, column=1).value or ""
            name_pattern = sheet.cell(row=2, column=1).value or ""
            
            # Headers are located on Row 4
            headers = []
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row=4, column=col).value
                headers.append(str(val) if val is not None else "")
            
            # Remove trailing empty headers
            while headers and headers[-1] == "":
                headers.pop()
                
            if not headers:
                continue
                
            # Data rows start from Row 5
            rows = []
            for r in range(5, sheet.max_row + 1):
                row_vals = {}
                has_value = False
                for col_idx, h in enumerate(headers):
                    cell_val = sheet.cell(row=r, column=col_idx + 1).value
                    if cell_val is not None:
                        has_value = True
                    row_vals[h] = str(cell_val) if cell_val is not None else ""
                if has_value:
                    rows.append(row_vals)
                    
            sheets_data[sheet_name] = {
                "template_path": template_path,
                "name_pattern": name_pattern,
                "headers": headers,
                "rows": rows
            }
        return sheets_data
    except Exception as e:
        st.error(f"Помилка при завантаженні конфігу {filepath}: {e}")
        return None

def save_excel_config(filepath, sheet_name, template_path, name_pattern, headers, df_data):
    """Saves changes back to Excel config safely preserving other sheets."""
    try:
        # Load the workbook preserving formula structures (don't use data_only=True)
        wb = openpyxl.load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return False
            
        sheet = wb[sheet_name]
        
        # 1. Update metadata in A1 and A2
        sheet.cell(row=1, column=1).value = template_path
        sheet.cell(row=2, column=1).value = name_pattern
        
        # 2. Update headers in Row 4
        for col_idx, h in enumerate(headers):
            sheet.cell(row=4, column=col_idx + 1).value = h
            
        # 3. Clear old data rows from Row 5 onwards
        max_r = max(sheet.max_row, 5)
        for r in range(5, max_r + 1):
            for c in range(1, len(headers) + 1):
                sheet.cell(row=r, column=c).value = None
                
        # 4. Write new data rows
        for r_idx, row_dict in enumerate(df_data):
            for c_idx, h in enumerate(headers):
                val = row_dict.get(h, "")
                cell = sheet.cell(row=5 + r_idx, column=c_idx + 1)
                
                # Check if it looks like a number
                if isinstance(val, str):
                    val_strip = val.strip()
                    if val_strip.isdigit():
                        cell.value = int(val_strip)
                    elif re.match(r'^\d+\.\d+$', val_strip):
                        cell.value = float(val_strip)
                    else:
                        cell.value = val
                else:
                    cell.value = val
                    
                # Enable multiline wrapping if newlines are present
                if isinstance(cell.value, str) and '\n' in cell.value:
                    from openpyxl.styles import Alignment
                    cell.alignment = Alignment(wrapText=True)
                    
        wb.save(filepath)
        return True
    except Exception as e:
        st.error(f"Помилка при збереженні змін: {e}")
        return False

# ----------------------------------------------------
# LIVE SUBPROCESS LOG STREAMING
# ----------------------------------------------------

def run_subprocess_and_stream(args):
    """Runs the python automation script and streams console outputs to Streamlit."""
    python_exe = sys.executable
    script_path = "_templates_machine_.py"
    
    cmd = [python_exe, script_path] + args
    
    log_area = st.empty()
    progress_bar = st.progress(0.0)
    
    st.info(f"🚀 Запущено команду: `python _templates_machine_.py {' '.join(f'\"{a}\"' for a in args)}`")
    
    # Initialize last operation state
    st.session_state["last_operation_logs"] = []
    st.session_state["last_operation_status"] = "running"
    st.session_state["last_operation_cmd"] = f"python _templates_machine_.py {' '.join(f'\"{a}\"' for a in args)}"
    
    # Force Python stdout to UTF-8 using environment variables
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    
    # Initialize process in binary mode to safely handle dynamic decoding fallbacks
    process = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=False,
        env=env
    )
    
    logs = []
    
    while True:
        line_bytes = process.stdout.readline()
        if not line_bytes:
            break
            
        # Bulletproof byte decoding fallback strategy (UTF-8 -> CP1251 -> UTF-8 with replacement)
        line = ""
        try:
            line = line_bytes.decode('utf-8')
        except UnicodeDecodeError:
            try:
                line = line_bytes.decode('cp1251')
            except Exception:
                line = line_bytes.decode('utf-8', errors='replace')
                
        cleaned_line = line.strip()
        if cleaned_line:
            logs.append(cleaned_line)
            st.session_state["last_operation_logs"] = logs
            # Display logs in real-time
            log_area.code("\n".join(logs[-15:])) # Show last 15 lines of progress
            
            # Simple progress heuristics
            if "Створено" in cleaned_line or "Обробка" in cleaned_line:
                progress_bar.progress(0.5)
            elif "Готово!" in cleaned_line or "Успішно!" in cleaned_line:
                progress_bar.progress(1.0)
                
    process.wait()
    
    if process.returncode == 0:
        progress_bar.progress(1.0)
        st.session_state["last_operation_status"] = "success"
        st.success("🎉 Завдання успішно виконано!")
        st.balloons()
    else:
        st.session_state["last_operation_status"] = "error"
        st.error(f"❌ Помилка виконання! Код завершення: {process.returncode}")
        
    return process.returncode, logs

# ----------------------------------------------------
# HELPER FOR PERSISTENT CONSOLE LOGS
# ----------------------------------------------------

def show_last_operation_logs():
    """Displays the persistent logs of the last executed operation from session state."""
    if "last_operation_logs" in st.session_state and st.session_state["last_operation_logs"]:
        status = st.session_state.get("last_operation_status", "")
        cmd = st.session_state.get("last_operation_cmd", "")
        
        st.markdown("---")
        if status == "success":
            st.markdown("### 🟢 Результат останньої операції: Успішно")
        elif status == "error":
            st.markdown("### 🔴 Результат останньої операції: Помилка")
        else:
            st.markdown("### 🟡 Результат останньої операції: Виконується")
            
        if cmd:
            st.caption(f"Команда: `{cmd}`")
            
        with st.expander("📋 Показати повний лог консолі", expanded=True):
            st.code("\n".join(st.session_state["last_operation_logs"]))

def move_autopilot_outputs(dest_dir):
    """Moves generated templates and configs of autopilot mode to a custom destination directory."""
    if not dest_dir:
        return
    try:
        import shutil
        dest_dir = os.path.abspath(dest_dir)
        os.makedirs(dest_dir, exist_ok=True)
        # Move Auto_Config.xlsx and Auto_Run_All.bat
        for filename in ["Auto_Config.xlsx", "Auto_Run_All.bat"]:
            src = os.path.join(os.getcwd(), filename)
            if os.path.exists(src):
                shutil.move(src, os.path.join(dest_dir, filename))
        # Move all template_* files from current directory
        moved_count = 0
        for filename in os.listdir(os.getcwd()):
            if filename.startswith("template_") and os.path.isfile(filename):
                shutil.move(filename, os.path.join(dest_dir, filename))
                moved_count += 1
        st.toast(f"✅ Результати аналізу перенесено в: {dest_dir} (переміщено {moved_count} шаблонів)", icon="📁")
    except Exception as e:
        st.error(f"Помилка при перенесенні файлів результатів: {e}")

def move_batch_outputs(sample_path, dest_dir):
    """Moves generated templates and configs of batch mode to a custom destination directory."""
    if not dest_dir or not sample_path:
        return
    try:
        import shutil
        dest_dir = os.path.abspath(dest_dir)
        os.makedirs(dest_dir, exist_ok=True)
        f_dir = os.path.dirname(os.path.abspath(sample_path))
        base_name, ext = os.path.splitext(os.path.basename(sample_path))
        
        files_to_move = [
            f"template_{base_name}{ext}",
            f"{base_name}_config.xlsx",
            f"{base_name}_run_all.bat"
        ]
        moved_count = 0
        for filename in files_to_move:
            src = os.path.join(f_dir, filename)
            if os.path.exists(src):
                shutil.move(src, os.path.join(dest_dir, filename))
                moved_count += 1
        st.toast(f"✅ Результати пакетного аналізу перенесено в: {dest_dir} (переміщено {moved_count} файлів)", icon="📁")
    except Exception as e:
        st.error(f"Помилка при перенесенні файлів результатів: {e}")

def move_pairwise_outputs(file1_path, dest_dir):
    """Moves generated templates and configs of pairwise mode to a custom destination directory."""
    if not dest_dir or not file1_path:
        return
    try:
        import shutil
        dest_dir = os.path.abspath(dest_dir)
        os.makedirs(dest_dir, exist_ok=True)
        f_dir = os.path.dirname(os.path.abspath(file1_path))
        base_name, ext = os.path.splitext(os.path.basename(file1_path))
        
        files_to_move = [
            f"template_{base_name}{ext}",
            f"{base_name}_config.xlsx"
        ]
        moved_count = 0
        for filename in files_to_move:
            src = os.path.join(f_dir, filename)
            if os.path.exists(src):
                shutil.move(src, os.path.join(dest_dir, filename))
                moved_count += 1
        st.toast(f"✅ Результати попарного порівняння перенесено в: {dest_dir} (переміщено {moved_count} файлів)", icon="📁")
    except Exception as e:
        st.error(f"Помилка при перенесенні файлів результатів: {e}")

# ----------------------------------------------------
# CORE DASHBOARD INTERFACE LAYOUT
# ----------------------------------------------------

# Initialize session state variables for file and folder picker widgets
for key in [
    "txt_auto_folder", "txt_batch_sample", "txt_batch_folder", 
    "txt_pair_file1", "txt_pair_file2", "editor_config_path", 
    "gen_config_path", "gen_output_dir", "analysis_output_dir",
    "last_operation_logs", "last_operation_status", "last_operation_cmd"
]:
    if key not in st.session_state:
        if key == "last_operation_logs":
            st.session_state[key] = []
        elif key in ["last_operation_status", "last_operation_cmd"]:
            st.session_state[key] = None if key == "last_operation_status" else ""
        else:
            st.session_state[key] = ""

configs, templates = scan_workspace()

# SIDEBAR: Workspace Metrics & Quick Guide
st.sidebar.markdown("""
<div style="text-align: center; margin-bottom: 2rem;">
    <span style="font-size: 3rem;">⚙️</span>
    <h3 style="margin-top: 0.5rem; font-weight: 700; color: #FF4B4B;">TemplateMachine</h3>
    <span style="color: #7f8c8d; font-size: 0.9rem;">Control Center v1.0</span>
</div>
""", unsafe_allow_html=True)

st.sidebar.subheader("📊 Стан робочого простору")
st.sidebar.metric("🗂️ Знайдено конфігів", len(configs))
st.sidebar.metric("📄 Знайдено шаблонів", len(templates))

st.sidebar.markdown("---")
st.sidebar.subheader("📖 Коротка довідка")
st.sidebar.info("""
**1. Аналіз:** Порівнює кілька файлів та створює розумний шаблон (`template_*`) та Excel-конфіг.
**2. Редагування:** Правити параметри генерації та значення змінних можна прямо в табі **Редактор конфігів**.
**3. Генерація:** Швидко генерує готові Word/Excel файли на основі обраного конфігу.
""")

st.sidebar.warning("""
⚠️ **Для безвіконного режиму (Headless):** Якщо додаток запущено на сервері без графічної оболонки, діалогові вікна вибору папок не відкриватимуться. Просто вводьте та копіюйте шляхи вручну в текстові поля!
""")

# MAIN PAGE HEADER
st.markdown('<div class="main-title">🚀 Панель керування TemplateMachine</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Універсальний комбайн для автоматизації документів та аналізу архівів</div>', unsafe_allow_html=True)

# TABS DEFINITION
tab_analysis, tab_editor, tab_generator, tab_help = st.tabs([
    "✈️ Аналіз та Створення Шаблонів",
    "📝 Редактор Excel Конфігів",
    "⚡ Генерація Документів",
    "📖 Повна Довідка"
])

# ----------------------------------------------------
# TAB 1: ARCHIVE ANALYSIS & TEMPLATE CREATION
# ----------------------------------------------------
with tab_analysis:
    st.header("🔍 Режими аналізу та розпізнавання шаблонів")
    st.write("Скрипт проведе інтелектуальне попарне або групове порівняння документів, виділить змінні і створить конфігураційний файл.")
    
    # Native Streamlit click callbacks to update widget states safely before instantiation
    def select_auto_folder():
        selected = open_folder_picker("Оберіть папку з архівом для аналізу")
        if selected:
            st.session_state["txt_auto_folder"] = selected
            
    def select_batch_sample():
        selected = open_file_picker()
        if selected:
            st.session_state["txt_batch_sample"] = selected
            
    def select_batch_folder():
        selected = open_folder_picker("Оберіть папку для пакетного аналізу")
        if selected:
            st.session_state["txt_batch_folder"] = selected
            
    def select_pair_file1():
        selected = open_file_picker()
        if selected:
            st.session_state["txt_pair_file1"] = selected
            
    def select_pair_file2():
        selected = open_file_picker()
        if selected:
            st.session_state["txt_pair_file2"] = selected
    
    mode = st.radio(
        "Оберіть режим аналізу:",
        [
            "✈️ Повний автопілот (Сканування папки та автоматичне розбиття на групи)",
            "🔍 Пакетний аналіз (Один файл-зразок + папка з іншими файлами)",
            "⚖️ Попарне порівняння (Точне порівняння двох конкретних файлів)"
        ],
        index=0
    )
    
    # --- ALWAYS-VISIBLE ANALYSIS OUTPUT DIRECTORY SELECTION ---
    st.markdown("##### 📁 Папка для збереження результатів (Необов'язково)")
    st.caption("Якщо вказано, створені шаблони (`template_*`), Excel-конфіги та `.bat` файли будуть перенесені в цю папку. У безвіконному (headless) режимі введіть шлях вручну.")
    
    col_ao1, col_ao2 = st.columns([3, 1])
    with col_ao1:
        a_out_dir = st.text_input(
            "📁 Шлях до папки результатів:",
            placeholder="Наприклад: C:/MyTemplates (залиште порожнім для збереження за замовчуванням)",
            key="analysis_output_dir"
        )
    with col_ao2:
        st.write(" ")
        st.write(" ")
        def select_analysis_output_dir():
            selected = open_folder_picker("Оберіть папку для збереження результатів")
            if selected:
                st.session_state["analysis_output_dir"] = selected
        st.button("📁 Обрати папку", key="btn_analysis_output_dir", on_click=select_analysis_output_dir)
        
    st.markdown("---")
    
    if "Повний автопілот" in mode:
        st.subheader("✈️ Режим 1: Повний автопілот")
        st.write("Аналізує вказану папку, групує схожі за структурою документи, створює для кожної групи окремий шаблон та зводить все в єдиний мульти-конфіг `Auto_Config.xlsx`.")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            folder_path = st.text_input(
                "Шлях до папки з архівом документів:",
                placeholder="Наприклад: example/docs",
                key="txt_auto_folder"
            )
        with col2:
            st.write(" ")
            st.write(" ")
            st.button("📁 Провідник Windows", key="btn_auto_folder", on_click=select_auto_folder)
                    
        if st.button("🚀 Запустити повний автопілот", type="primary"):
            if not folder_path:
                st.error("Будь ласка, вкажіть шлях до папки!")
            elif not os.path.exists(folder_path):
                st.error(f"Вказана папка '{folder_path}' не існує!")
            else:
                ret_code, _ = run_subprocess_and_stream([folder_path])
                if ret_code == 0 and st.session_state.get("analysis_output_dir"):
                    move_autopilot_outputs(st.session_state["analysis_output_dir"])
                configs, templates = scan_workspace()
                st.rerun()
                
    elif "Пакетний аналіз" in mode:
        st.subheader("🔍 Режим 2: Пакетний аналіз")
        st.write("Порівнює один обраний файл-зразок зі схожими файлами у папці та формує шаблон і конфігурацію.")
        
        # Row for Sample file
        col1, col2 = st.columns([3, 1])
        with col1:
            sample_file = st.text_input(
                "Шлях до файлу-зразка (.docx або .xlsx):",
                placeholder="Наприклад: example/w.docx",
                key="txt_batch_sample"
            )
        with col2:
            st.write(" ")
            st.write(" ")
            st.button("📄 Обрати зразок", key="btn_batch_sample", on_click=select_batch_sample)
                    
        # Row for Folder
        col1, col2 = st.columns([3, 1])
        with col1:
            folder_path = st.text_input(
                "Шлях до папки порівняння:",
                placeholder="Наприклад: example",
                key="txt_batch_folder"
            )
        with col2:
            st.write(" ")
            st.write(" ")
            st.button("📁 Обрати папку", key="btn_batch_folder", on_click=select_batch_folder)
                    
        if st.button("🚀 Запустити пакетний аналіз", type="primary"):
            if not sample_file or not folder_path:
                st.error("Будь ласка, вкажіть і файл-зразок, і папку порівняння!")
            elif not os.path.exists(sample_file):
                st.error(f"Файл-зразок '{sample_file}' не знайдено!")
            elif not os.path.exists(folder_path):
                st.error(f"Папку порівняння '{folder_path}' не знайдено!")
            else:
                ret_code, _ = run_subprocess_and_stream([sample_file, folder_path])
                if ret_code == 0 and st.session_state.get("analysis_output_dir"):
                    move_batch_outputs(sample_file, st.session_state["analysis_output_dir"])
                configs, templates = scan_workspace()
                st.rerun()
                
    elif "Попарне порівняння" in mode:
        st.subheader("⚖️ Режим 3: Попарне порівняння")
        st.write("Знаходить розбіжності між двома файлами, виділяє змінні та створює шаблон та індивідуальний конфіг.")
        
        # File 1
        col1, col2 = st.columns([3, 1])
        with col1:
            file_1 = st.text_input(
                "Шлях до Першого файлу:",
                placeholder="Наприклад: example/w1.docx",
                key="txt_pair_file1"
            )
        with col2:
            st.write(" ")
            st.write(" ")
            st.button("📄 Обрати файл 1", key="btn_pair_file1", on_click=select_pair_file1)
                    
        # File 2
        col1, col2 = st.columns([3, 1])
        with col1:
            file_2 = st.text_input(
                "Шлях до Другого файлу:",
                placeholder="Наприклад: example/w2.docx",
                key="txt_pair_file2"
            )
        with col2:
            st.write(" ")
            st.write(" ")
            st.button("📄 Обрати файл 2", key="btn_pair_file2", on_click=select_pair_file2)
                    
        if st.button("🚀 Запустити попарне порівняння", type="primary"):
            if not file_1 or not file_2:
                st.error("Будь ласка, вкажіть обидва файли для порівняння!")
            elif not os.path.exists(file_1) or not os.path.exists(file_2):
                st.error("Один або обидва вказані файли не знайдені!")
            else:
                ret_code, _ = run_subprocess_and_stream([file_1, file_2])
                if ret_code == 0 and st.session_state.get("analysis_output_dir"):
                    move_pairwise_outputs(file_1, st.session_state["analysis_output_dir"])
                configs, templates = scan_workspace()
                st.rerun()

    # Persistent log viewer at the bottom of analysis tab
    show_last_operation_logs()

# ----------------------------------------------------
# TAB 2: INTERACTIVE EXCEL CONFIG EDITOR
# ----------------------------------------------------
with tab_editor:
    st.header("📝 Інтерактивний редактор файлів конфігурації")
    st.write("Оберіть будь-який створений конфіг-файл, щоб відредагувати параметри шаблону, імені або змінити дані змінних та аркушів.")
    
    col_c1, col_c2 = st.columns([3, 1])
    with col_c1:
        selected_config = st.text_input(
            "Шлях до файлу конфігурації для редагування:",
            placeholder="Оберіть Excel-файл конфігурації...",
            key="editor_config_path"
        )
    with col_c2:
        st.write(" ")
        st.write(" ")
        def select_editor_config():
            selected = open_file_picker(filetypes=[("Excel конфігурації", "*.xlsx"), ("Усі файли", "*.*")])
            if selected:
                st.session_state["editor_config_path"] = selected
        st.button("📁 Обрати конфіг", key="btn_editor_config", on_click=select_editor_config)
        
    if configs:
        quick_select = st.selectbox(
            "Або оберіть зі знайдених у робочому просторі:",
            [""] + configs,
            key="quick_editor_config_select"
        )
        if quick_select:
            st.session_state["editor_config_path"] = quick_select
            
    cfg_path = st.session_state["editor_config_path"]
    if not cfg_path:
        st.info("Будь ласка, оберіть файл конфігурації Excel (`*.xlsx`) для початку роботи!")
    elif not os.path.exists(cfg_path):
        st.error(f"Вказаний файл конфігурації '{cfg_path}' не знайдено!")
    else:
        sheets_data = load_excel_config(cfg_path)
        
        if sheets_data:
            sheet_names = list(sheets_data.keys())
            
            if "editor_selected_sheet" not in st.session_state or st.session_state["editor_selected_sheet"] not in sheet_names:
                st.session_state["editor_selected_sheet"] = sheet_names[0] if sheet_names else ""
                
            selected_sheet = st.selectbox(
                "Оберіть аркуш конфігурації:",
                sheet_names,
                index=sheet_names.index(st.session_state["editor_selected_sheet"]) if st.session_state["editor_selected_sheet"] in sheet_names else 0,
                key="widget_editor_sheet"
            )
            
            if selected_sheet != st.session_state["editor_selected_sheet"]:
                st.session_state["editor_selected_sheet"] = selected_sheet
                
            # --- SHEET CRUD MANAGEMENT ---
            with st.expander("📂 Керування аркушами (Sheets Operations)", expanded=False):
                col_s1, col_s2, col_s3 = st.columns(3)
                
                with col_s1:
                    st.markdown("##### ➕ Створити новий аркуш")
                    new_s_name = st.text_input("Назва нового аркуша:", key="txt_new_sheet_name")
                    if st.button("➕ Створити аркуш", key="btn_create_sheet"):
                        if not new_s_name:
                            st.error("Введіть назву аркуша!")
                        else:
                            clean_s_name = re.sub(r'[\\/*?:\[\]]', "", new_s_name)[:31].strip()
                            if create_new_sheet(cfg_path, clean_s_name):
                                st.session_state["editor_selected_sheet"] = clean_s_name
                                st.success(f"Аркуш '{clean_s_name}' успішно створено!")
                                time.sleep(1)
                                st.rerun()
                                
                with col_s2:
                    st.markdown("##### ✏️ Перейменувати поточний аркуш")
                    new_rename_name = st.text_input("Нова назва аркуша:", key="txt_rename_sheet_name")
                    if st.button("✏️ Перейменувати аркуш", key="btn_rename_sheet"):
                        if not new_rename_name:
                            st.error("Введіть нову назву!")
                        else:
                            clean_rename_name = re.sub(r'[\\/*?:\[\]]', "", new_rename_name)[:31].strip()
                            if rename_sheet(cfg_path, selected_sheet, clean_rename_name):
                                st.session_state["editor_selected_sheet"] = clean_rename_name
                                st.success(f"Аркуш перейменовано на '{clean_rename_name}'!")
                                time.sleep(1)
                                st.rerun()
                                
                with col_s3:
                    st.markdown("##### ❌ Видалити поточний аркуш")
                    st.warning("Ця дія є незворотною!")
                    confirm_delete = st.checkbox("Підтверджую видалення аркуша", key="chk_confirm_delete_sheet")
                    if st.button("❌ Видалити аркуш", key="btn_delete_sheet"):
                        if not confirm_delete:
                            st.error("Будь ласка, підтвердіть видалення чекбоксом!")
                        else:
                            if delete_sheet(cfg_path, selected_sheet):
                                st.success(f"Аркуш '{selected_sheet}' видалено!")
                                remaining_sheets = [s for s in sheet_names if s != selected_sheet]
                                st.session_state["editor_selected_sheet"] = remaining_sheets[0] if remaining_sheets else ""
                                time.sleep(1)
                                st.rerun()
                                
            # --- LOAD SHEET DATA ---
            sheet_info = sheets_data[selected_sheet]
            config_sheet_key = f"{cfg_path}_{selected_sheet}"
            
            # State synchronization on config/sheet change
            if st.session_state.get("loaded_config_sheet") != config_sheet_key:
                st.session_state["loaded_config_sheet"] = config_sheet_key
                st.session_state["current_sheet_headers"] = list(sheet_info["headers"])
                st.session_state["current_sheet_data"] = list(sheet_info["rows"])
                st.session_state["editor_template_path"] = sheet_info["template_path"]
                st.session_state["editor_name_pattern"] = sheet_info["name_pattern"]
                
            # --- METADATA (A1 & A2) ---
            st.markdown("### ⚙️ Налаштування генерації для обраного аркуша")
            col1, col2 = st.columns(2)
            
            with col1:
                col_t1, col_t2 = st.columns([3, 1])
                with col_t1:
                    t_path = st.text_input(
                        "📄 Шлях до шаблону (комірка A1):",
                        placeholder="Оберіть файл шаблону...",
                        key="editor_template_path"
                    )
                with col_t2:
                    st.write(" ")
                    st.write(" ")
                    def select_editor_template():
                        selected = open_file_picker(filetypes=[
                            ("Усі шаблони", "*.docx;*.xlsx"),
                            ("Word шаблони", "*.docx"),
                            ("Excel шаблони", "*.xlsx"),
                            ("Усі файли", "*.*")
                        ])
                        if selected:
                            st.session_state["editor_template_path"] = selected
                    st.button("📁 Шаблон", key="btn_editor_template", on_click=select_editor_template)
                    
            with col2:
                n_pattern = st.text_input(
                    "✍️ Шаблон імені вихідних файлів (комірка A2):",
                    placeholder="Наприклад: output_{{YYYY}}.docx",
                    key="editor_name_pattern"
                )
                
            # --- VARIABLE & COLUMN EDITOR ---
            st.markdown("### 🛠️ Керування змінними (колонками) та структурою")
            
            if st.session_state["current_sheet_headers"]:
                st.markdown("##### 🏷️ Наявні змінні на цьому аркуші:")
                badges_html = " ".join([
                    f'<span style="background-color: #2e86de; color: white; padding: 4px 10px; margin: 4px; border-radius: 12px; display: inline-block; font-size: 0.85rem; font-weight: 600; box-shadow: 0 1px 3px rgba(0,0,0,0.15);">{h}</span>'
                    for h in st.session_state["current_sheet_headers"]
                ])
                st.markdown(badges_html, unsafe_allow_html=True)
            else:
                st.info("Немає активних змінних. Створіть першу змінну нижче.")
                
            st.write(" ")
            col_v1, col_v2, col_v3 = st.columns(3)
            
            with col_v1:
                st.markdown("##### ➕ Додати нову змінну")
                new_var_name = st.text_input("Ім'я змінної (напр., client_name):", key="txt_new_var_name")
                if st.button("➕ Додати змінну", key="btn_add_var"):
                    if not new_var_name:
                        st.error("Введіть ім'я змінної!")
                    elif new_var_name in st.session_state["current_sheet_headers"]:
                        st.error("Така змінна вже існує!")
                    else:
                        st.session_state["current_sheet_headers"].append(new_var_name)
                        for row in st.session_state["current_sheet_data"]:
                            row[new_var_name] = ""
                        st.success(f"Змінну '{new_var_name}' додано!")
                        st.rerun()
                        
            with col_v2:
                st.markdown("##### ✏️ Перейменувати змінну")
                if st.session_state["current_sheet_headers"]:
                    var_to_rename = st.selectbox("Оберіть змінну:", st.session_state["current_sheet_headers"], key="sel_rename_var")
                    rename_var_name = st.text_input("Нове ім'я змінної:", key="txt_rename_var_name")
                    if st.button("✏️ Перейменувати змінну", key="btn_rename_var"):
                        if not rename_var_name:
                            st.error("Введіть нове ім'я!")
                        elif rename_var_name in st.session_state["current_sheet_headers"]:
                            st.error("Змінна з таким ім'ям вже існує!")
                        else:
                            h_idx = st.session_state["current_sheet_headers"].index(var_to_rename)
                            st.session_state["current_sheet_headers"][h_idx] = rename_var_name
                            for row in st.session_state["current_sheet_data"]:
                                if var_to_rename in row:
                                    row[rename_var_name] = row.pop(var_to_rename)
                            st.success(f"Змінну перейменовано на '{rename_var_name}'!")
                            st.rerun()
                else:
                    st.caption("Немає доступних змінних.")
                    
            with col_v3:
                st.markdown("##### ❌ Видалити змінну")
                if st.session_state["current_sheet_headers"]:
                    var_to_delete = st.selectbox("Оберіть змінну для видалення:", st.session_state["current_sheet_headers"], key="sel_delete_var")
                    confirm_var_delete = st.checkbox("Підтверджую видалення", key="chk_confirm_delete_var")
                    if st.button("❌ Видалити змінну", key="btn_delete_var"):
                        if not confirm_var_delete:
                            st.error("Підтвердіть видалення чекбоксом!")
                        else:
                            st.session_state["current_sheet_headers"].remove(var_to_delete)
                            for row in st.session_state["current_sheet_data"]:
                                if var_to_delete in row:
                                    row.pop(var_to_delete)
                            st.success(f"Змінну '{var_to_delete}' видалено!")
                            st.rerun()
                else:
                    st.caption("Немає доступних змінних.")
                    
            # --- DATA EDITOR AND SAVE ---
            st.markdown("### 📊 Дані рядків для генерації документів (починаючи з рядка 5)")
            st.caption("Клікніть двічі для редагування. Кнопка '+' внизу таблиці додає порожній рядок.")
            
            headers = st.session_state["current_sheet_headers"]
            rows = st.session_state["current_sheet_data"]
            
            if rows:
                df = pd.DataFrame(rows, columns=headers)
            else:
                df = pd.DataFrame(columns=headers)
                
            edited_df = st.data_editor(
                df,
                num_rows="dynamic",
                use_container_width=True,
                key="config_data_editor"
            )
            
            st.write(" ")
            
            col_save1, col_save2 = st.columns([1, 4])
            with col_save1:
                if st.button("💾 Зберегти зміни в Excel", type="primary", key="btn_save_config"):
                    clean_rows = []
                    for _, r in edited_df.iterrows():
                        row_dict = {}
                        for h in headers:
                            val = r.get(h, "")
                            row_dict[h] = str(val) if pd.notna(val) else ""
                        clean_rows.append(row_dict)
                        
                    st.session_state["current_sheet_data"] = clean_rows
                    
                    success = save_excel_config(
                        cfg_path,
                        selected_sheet,
                        st.session_state["editor_template_path"],
                        st.session_state["editor_name_pattern"],
                        headers,
                        clean_rows
                    )
                    if success:
                        st.success("🎉 Всі зміни успішно записані в Excel файл!")
                        st.balloons()
                        time.sleep(1)
                        st.rerun()

# ----------------------------------------------------
# TAB 3: DOCUMENT GENERATION & PROGRESS TRACKING
# ----------------------------------------------------
with tab_generator:
    st.header("⚡ Масова генерація документів")
    st.write("Оберіть файл конфігурації, перевірте шлях до шаблону і запустіть автоматичне заповнення.")
    
    col_g1, col_g2 = st.columns([3, 1])
    with col_g1:
        g_config_input = st.text_input(
            "Шлях до файлу конфігурації для генерації:",
            placeholder="Оберіть Excel-файл конфігурації...",
            key="gen_config_path"
        )
    with col_g2:
        st.write(" ")
        st.write(" ")
        def select_gen_config():
            selected = open_file_picker(filetypes=[("Excel конфігурації", "*.xlsx"), ("Усі файли", "*.*")])
            if selected:
                st.session_state["gen_config_path"] = selected
        st.button("📁 Обрати конфіг", key="btn_gen_config", on_click=select_gen_config)
        
    if configs:
        quick_gen_select = st.selectbox(
            "Або оберіть зі знайдених у робочому просторі:",
            [""] + configs,
            key="quick_gen_config_select"
        )
        if quick_gen_select:
            st.session_state["gen_config_path"] = quick_gen_select
            
    g_path = st.session_state["gen_config_path"]
    
    # --- ALWAYS-VISIBLE OUTPUT DIRECTORY SELECTION ---
    st.markdown("##### 📁 Папка для збереження готових документів (Необов'язково)")
    st.caption("Якщо не вказано, файли зберігатимуться відносно папки Excel-конфігурації. У безвіконному (headless) режимі введіть шлях вручну.")
    
    col_go1, col_go2 = st.columns([3, 1])
    with col_go1:
        g_out_dir = st.text_input(
            "📁 Шлях до папки збереження:",
            placeholder="Наприклад: C:/ProcessedDocs (залиште порожнім за замовчуванням)",
            key="gen_output_dir"
        )
    with col_go2:
        st.write(" ")
        st.write(" ")
        def select_gen_output_dir():
            selected = open_folder_picker("Оберіть папку для збереження готових документів")
            if selected:
                st.session_state["gen_output_dir"] = selected
        st.button("📁 Обрати папку", key="btn_gen_output_dir", on_click=select_gen_output_dir)
    if not g_path:
        st.info("Будь ласка, оберіть Excel-файл конфігурації для генерації!")
    elif not os.path.exists(g_path):
        st.error(f"Вказаний файл конфігурації '{g_path}' не знайдено!")
    else:
        g_sheets_data = load_excel_config(g_path)
        
        if g_sheets_data:
            g_sheet_names = ["all (Всі аркуші)"] + list(g_sheets_data.keys())
            selected_g_sheet = st.selectbox("Оберіть аркуш для обробки:", g_sheet_names, key="gen_sheet_select")
            
            actual_sheet_name = ""
            if selected_g_sheet != "all (Всі аркуші)":
                actual_sheet_name = selected_g_sheet
                
            # State synchronization on config/sheet change for generator
            gen_sheet_key = f"gen_{g_path}_{actual_sheet_name}"
            if st.session_state.get("loaded_gen_sheet") != gen_sheet_key:
                st.session_state["loaded_gen_sheet"] = gen_sheet_key
                if actual_sheet_name:
                    st.session_state["gen_template_path"] = g_sheets_data[actual_sheet_name]["template_path"]
                else:
                    st.session_state["gen_template_path"] = ""
                
            # --- VIEW/CHANGE TEMPLATE FILE VIA DIALOG ---
            if actual_sheet_name:
                sheet_data = g_sheets_data[actual_sheet_name]
                current_template = sheet_data["template_path"]
                
                st.markdown("##### 📄 Перевірка та зміна файлу шаблону (комірка A1)")
                
                col_gt1, col_gt2 = st.columns([3, 1])
                with col_gt1:
                    new_gt_path = st.text_input(
                        "📄 Поточний шлях до шаблону:",
                        placeholder="Оберіть файл шаблону...",
                        key="gen_template_path"
                    )
                with col_gt2:
                    st.write(" ")
                    st.write(" ")
                    def select_gen_template():
                        selected = open_file_picker(filetypes=[
                            ("Усі шаблони", "*.docx;*.xlsx"),
                            ("Word шаблони", "*.docx"),
                            ("Excel шаблони", "*.xlsx"),
                            ("Усі файли", "*.*")
                        ])
                        if selected:
                            st.session_state["gen_template_path"] = selected
                            if update_config_template_path(g_path, actual_sheet_name, selected):
                                st.toast(f"✅ Шаблон оновлено в Excel: {os.path.basename(selected)}", icon="💾")
                    st.button("📁 Змінити шаблон", key="btn_gen_template", on_click=select_gen_template)
                    
                # If path was changed manually
                if st.session_state.get("gen_template_path", "") != current_template:
                    if st.button("💾 Зберегти новий шлях шаблону в Excel", key="btn_save_gen_template"):
                        manually_entered = st.session_state.get("gen_template_path", "")
                        if update_config_template_path(g_path, actual_sheet_name, manually_entered):
                            st.success("Шлях шаблону успішно оновлено в конфігурації!")
                            time.sleep(1)
                            st.rerun()
                            
            # Row selector filtering options
            selected_g_row = "all"
            if selected_g_sheet != "all (Всі аркуші)":
                rows_count = len(sheet_data["rows"])
                
                row_options = ["all (Всі рядки з даними)"]
                for i in range(rows_count):
                    preview_fields = []
                    row_dict = sheet_data["rows"][i]
                    headers_to_show = sheet_data["headers"][:3]
                    for h in headers_to_show:
                        val = row_dict.get(h, "")
                        if val:
                            preview_fields.append(f"{h}: {val}")
                    preview_str = ", ".join(preview_fields) if preview_fields else "Порожній рядок"
                    row_options.append(f"{5 + i} — ({preview_str})")
                    
                selected_g_row_str = st.selectbox("Оберіть рядок для обробки (за номером рядка в Excel):", row_options, key="gen_row_select")
                
                if "all" in selected_g_row_str:
                    selected_g_row = "all"
                else:
                    selected_g_row = selected_g_row_str.split(" — ")[0].strip()
                    
            st.markdown("---")
            st.subheader("🎯 Параметри запуску")
            
            st.markdown(f"""
            - **Файл конфігурації:** `{g_path}`
            - **Аркуш:** `{selected_g_sheet.split(' (')[0]}`
            - **Рядок:** `{selected_g_row}`
            - **Папка збереження:** `{st.session_state["gen_output_dir"] if st.session_state["gen_output_dir"] else "Папка конфігурації (за замовчуванням)"}`
            """)
            
            if st.button("⚡ Запустити генерацію документів", type="primary", key="btn_run_generation"):
                args = [g_path]
                
                sheet_arg = selected_g_sheet.split(' (')[0]
                row_arg = selected_g_row
                
                if sheet_arg != "all":
                    args.append(sheet_arg)
                    if row_arg != "all":
                        args.append(row_arg)
                elif row_arg != "all":
                    args.append("all")
                    args.append(row_arg)
                    
                # Append custom output folder if provided
                if st.session_state["gen_output_dir"]:
                    while len(args) < 3:
                        args.append("all")
                    args.append(st.session_state["gen_output_dir"])
                    
                run_subprocess_and_stream(args)
                st.rerun() # Refresh to display persistent logs immediately

    # Persistent log viewer at the bottom of generator tab
    show_last_operation_logs()

# ----------------------------------------------------
# TAB 4: HELP & DOCUMENTATION
# ----------------------------------------------------
with tab_help:
    st.header("📖 Повний посібник користувача.")
    st.write("Детальний опис можливостей та технічний посібник роботи комбайна (завантажено з _templates_machine_.txt).")
    
    st.markdown("---")
    
    doc_markdown = get_formatted_documentation_markdown()
    st.markdown(doc_markdown, unsafe_allow_html=True)
