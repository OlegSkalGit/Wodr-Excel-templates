import sys
import os
import re
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
from docxtpl import DocxTemplate, RichText, Listing

def clean_path_segment(segment):
    # Remove characters invalid for Windows filenames
    return re.sub(r'[*?:"<>|]', "", segment).strip()

def get_unique_path(full_path):
    directory = os.path.dirname(full_path)
    filename = os.path.basename(full_path)
    name, ext = os.path.splitext(filename)
    if not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)
    counter = 1
    unique_path = full_path
    while os.path.exists(unique_path):
        unique_path = os.path.join(directory, f"{name}_{counter}{ext}")
        counter += 1
    return unique_path

def render_string_template(template_str, variables):
    result = template_str
    for key, val in variables.items():
        # Match {{key}} with optional spaces: {{ key }}
        pattern = r"\{\{\s*" + re.escape(key) + r"\s*\}\}"
        result = re.sub(pattern, str(val), result)
    return result

def get_now_vars():
    now = datetime.now()
    return {
        "YYYY": now.strftime("%Y"),
        "MM": now.strftime("%m"),
        "DD": now.strftime("%d"),
        "hh": now.strftime("%H"),
        "mm": now.strftime("%M"),
        "ss": now.strftime("%S")
    }

def process_word(template_path, output_path, variables):
    try:
        doc = DocxTemplate(template_path)
        processed_vars = {}
        for key, value in variables.items():
            # Handle multiline - use Listing to preserve placeholder formatting
            if isinstance(value, str) and '\n' in value:
                processed_vars[key] = Listing(value)
            else:
                processed_vars[key] = value
        doc.render(processed_vars)
        doc.save(output_path)
        print(f"  [Word] Создан: {os.path.basename(output_path)}")
    except Exception as e:
        print(f"  [Ошибка Word]: {e}")

def process_excel(template_path, output_path, variables):
    try:
        wb = openpyxl.load_workbook(template_path)
        # Prepare string versions for substitution
        str_vars = {k: str(v) for k, v in variables.items()}
        
        for sheet in wb.worksheets:
            # 0. Handle sheet title replacement
            new_title = sheet.title
            title_modified = False
            for key, val in str_vars.items():
                pattern = r"\{\{\s*" + re.escape(key) + r"\s*\}\}"
                if re.search(pattern, new_title):
                    new_title = re.sub(pattern, val, new_title)
                    title_modified = True
            
            if title_modified:
                # Excel limits: max 31 chars, no /\?*:[]
                clean_title = re.sub(r'[\\/*?:\[\]]', "", new_title)[:31].strip()
                if not clean_title:
                    clean_title = "Sheet"
                
                # Ensure uniqueness within the workbook
                base_title = clean_title
                counter = 1
                while clean_title in wb.sheetnames and clean_title != sheet.title:
                    suffix = f"_{counter}"
                    # Adjust base to fit suffix
                    limit = 31 - len(suffix)
                    clean_title = base_title[:limit] + suffix
                    counter += 1
                
                sheet.title = clean_title

            # 1. Process cells
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        
                        # 1. Try exact match to preserve data type (e.g. number, date)
                        exact_key = None
                        for key in variables.keys():
                            # Pattern for exact match: {{ key }} or {{key}}
                            pattern = r"^\s*\{\{\s*" + re.escape(key) + r"\s*\}\}\s*$"
                            if re.match(pattern, original_value):
                                exact_key = key
                                break
                        
                        if exact_key is not None:
                            new_value = variables[exact_key]
                            source_type = variables.get(f"{exact_key}_type_code", "s")
                            template_type = cell.data_type
                            
                            # Decide on target type (source type priority)
                            target_type = source_type if source_type != "s" else template_type
                            
                            # If we have a string but target is number/date, try to cast
                            if isinstance(new_value, str):
                                if target_type == "n":
                                    try:
                                        if "." in new_value or "e" in new_value.lower():
                                            new_value = float(new_value)
                                        else:
                                            new_value = int(new_value)
                                    except:
                                        pass
                                elif target_type == "d":
                                    try:
                                        new_value = datetime.fromisoformat(new_value)
                                    except:
                                        pass
                                elif target_type == "b":
                                    if new_value.lower() in ["true", "1", "yes"]: new_value = True
                                    elif new_value.lower() in ["false", "0", "no"]: new_value = False
                            
                            cell.value = new_value
                            
                            if isinstance(cell.value, str) and '\n' in cell.value:
                                cell.alignment = Alignment(wrapText=True)
                            continue

                        # 2. Otherwise do partial string substitution
                        new_value = original_value
                        modified = False
                        for key, val in str_vars.items():
                            pattern = r"\{\{\s*" + re.escape(key) + r"\s*\}\}"
                            if re.search(pattern, new_value):
                                new_value = re.sub(pattern, val, new_value)
                                modified = True
                        
                        if modified:
                            cell.value = new_value
                            if isinstance(new_value, str) and '\n' in new_value:
                                cell.alignment = Alignment(wrapText=True)
        wb.save(output_path)
        print(f"  [Excel] Создан: {os.path.basename(output_path)}")
    except Exception as e:
        print(f"  [Ошибка Excel]: {e}")

def resolve_path(base_dir, path):
    if os.path.isabs(path):
        return path
    # Replace forward slashes with backslashes for consistency if on Windows
    # though Python handles both.
    return os.path.abspath(os.path.join(base_dir, path))

def main():
    if len(sys.argv) < 2:
        print("Использование: run.bat <путь_к_excel> [лист: all/номер] [рядок: all/номер]")
        return

    excel_file = sys.argv[1]
    sheet_selector = sys.argv[2] if len(sys.argv) > 2 else "all"
    row_selector = sys.argv[3] if len(sys.argv) > 3 else "all"

    if not os.path.exists(excel_file):
        print(f"Ошибка: Файл {excel_file} не найден.")
        return

    excel_dir = os.path.dirname(os.path.abspath(excel_file))
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        print(f"Ошибка при открытии Excel: {e}")
        return

    now_vars = get_now_vars()
    # Add types for system variables
    for k in list(now_vars.keys()):
        now_vars[f"{k}_type"] = "s"
    
    # Mapping for Excel data types to readable names
    TYPE_MAP = {
        's': 'string',
        'n': 'number',
        'd': 'date',
        'b': 'boolean',
        'f': 'formula',
        'e': 'error'
    }
    # Identify sheets to process
    sheets_to_process = []
    if sheet_selector.lower() == "all":
        sheets_to_process = wb.worksheets
    else:
        try:
            # Try as 1-based index
            idx = int(sheet_selector) - 1
            if 0 <= idx < len(wb.worksheets):
                sheets_to_process = [wb.worksheets[idx]]
        except ValueError:
            # Try as name
            if sheet_selector in wb.sheetnames:
                sheets_to_process = [wb[sheet_selector]]
    
    if not sheets_to_process:
        print(f"Лист '{sheet_selector}' не найден.")
        return

    for sheet in sheets_to_process:
        print(f"Обработка листа: {sheet.title}")
        
        # Row 1: Template Name
        template_rel_path = sheet.cell(row=1, column=1).value
        if not template_rel_path:
            print(f"  Пропуск листа {sheet.title}: Не указан шаблон в ячейке A1.")
            continue
        
        template_path = resolve_path(excel_dir, str(template_rel_path))
        if not os.path.exists(template_path):
            print(f"  Ошибка: Шаблон не найден: {template_path}")
            continue

        # Row 2: Generated files name pattern
        output_pattern = sheet.cell(row=2, column=1).value
        if not output_pattern:
            print(f"  Пропуск листа {sheet.title}: Не указан путь результата в ячейке A2.")
            continue

        # Row 4: Headers
        headers = []
        # Read headers from row 4, column 1 to max_column
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=4, column=col).value
            if val is not None:
                header_name = str(val).strip()
                # If user accidentally included {{ }} in the header cell, strip them
                if header_name.startswith("{{") and header_name.endswith("}}"):
                    header_name = header_name[2:-2].strip()
                headers.append(header_name)
            else:
                # If we hit an empty cell, we continue to check if there are more headers later
                # but usually they are contiguous. Let's just store None to keep indices aligned if needed,
                # or skip. Here we skip but keep track of the column index for data.
                headers.append(None)
        
        # Filter out trailing Nones and check if we have any headers
        while headers and headers[-1] is None:
            headers.pop()
            
        if not any(h is not None for h in headers):
            print(f"  Пропуск листа {sheet.title}: Не найдены заголовки в строке 4.")
            continue
        
        print(f"  Найденные переменные в Excel: {', '.join([h for h in headers if h is not None])}")

        # Row 5+: Data
        data_rows = []
        max_row = sheet.max_row
        
        # Row selector logic
        if row_selector.lower() == "all":
            for r in range(5, max_row + 1):
                row_vals = []
                for c in range(1, len(headers) + 1):
                    cell = sheet.cell(row=r, column=c)
                    row_vals.append((cell.value, cell.data_type))
                # Check if row is not empty
                if any(v[0] is not None for v in row_vals):
                    data_rows.append((r, row_vals))
        else:
            try:
                # Row selector as absolute Excel row number (>=5) or 1-based relative to data
                r_idx = int(row_selector)
                
                if 5 <= r_idx <= max_row:
                    row_vals = []
                    for c in range(1, len(headers) + 1):
                        cell = sheet.cell(row=r_idx, column=c)
                        row_vals.append((cell.value, cell.data_type))
                    if any(v[0] is not None for v in row_vals):
                        data_rows.append((r_idx, row_vals))
            except ValueError:
                print(f"  Неверный формат номера строки: {row_selector}")

        for r_num, row_vals in data_rows:
            variables = {**now_vars}
            for i, h in enumerate(headers):
                if h is None: continue
                val, v_type = row_vals[i] if i < len(row_vals) else (None, "s")
                
                # Keep original type for cell-to-cell transfer, 
                # but handle None as empty string for Word/Path templates
                if val is None:
                    val = ""
                
                variables[h] = val
                variables[f"{h}_type"] = TYPE_MAP.get(v_type, v_type)
                variables[f"{h}_type_code"] = v_type
            
            # Generate output path
            rendered_out_path = render_string_template(str(output_pattern), variables)
            
            # Ensure extension
            ext = os.path.splitext(template_path)[1].lower()
            if not rendered_out_path.lower().endswith(ext):
                rendered_out_path += ext
            
            # Fix separators and resolve
            rendered_out_path = rendered_out_path.replace('/', os.sep).replace('\\', os.sep)
            final_output_path = resolve_path(excel_dir, rendered_out_path)
            
            # Unique path to avoid overwriting
            final_output_path = get_unique_path(final_output_path)
            
            print(f"  Рядок {r_num}:")
            if ext == '.docx':
                process_word(template_path, final_output_path, variables)
            elif ext == '.xlsx':
                process_excel(template_path, final_output_path, variables)
            else:
                print(f"  [Ошибка] Неподдерживаемый формат шаблона: {ext}")

if __name__ == "__main__":
    main()
